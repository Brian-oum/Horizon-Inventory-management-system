import csv
from django.utils.dateparse import parse_date
from .forms import (
    CustomCreationForm, OEMForm, DeviceForm, DeviceRequestForm
)
from .models import (
    Device, OEM, DeviceRequest, Client, IssuanceRecord, ReturnRecord, Branch, Profile, DeviceSelection, DeviceIMEI,
    SelectedDevice
)
from django.contrib.auth.models import User
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth import login, logout
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
from django.db.models import Q, F, Count, Sum, Value, IntegerField
from django.db import transaction
from django.core.mail import send_mail
from django.http import HttpResponse, JsonResponse
from django.urls import reverse
from django.core.paginator import Paginator
from django.views.decorators.http import require_POST
from collections import defaultdict
import openpyxl
from openpyxl.utils import get_column_letter
import logging
from django.db.models.functions import Coalesce
from .models import PurchaseOrder
from .forms import PurchaseOrderForm
from django.utils import timezone
from .models import DeviceSelectionGroup  # add import at top
logger = logging.getLogger(__name__)


def custom_login(request):
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            if user.is_staff:
                return redirect('store_clerk_dashboard')
            else:
                return redirect('requestor_dashboard')
        else:
            messages.error(request, "Invalid username or password.")
    else:
        form = AuthenticationForm()
    return render(request, 'invent/login.html', {'form': form})


def logout_view(request):
    logout(request)
    messages.info(request, "You have been logged out.")
    return redirect('login')

# --- Dashboard for Requestor ---


@login_required
def requestor_dashboard(request):
    user_requests = DeviceRequest.objects.filter(
        requestor=request.user).order_by('id')
    labeled_requests = []
    total_requests_count = user_requests.count()
    for idx, req in enumerate(user_requests, start=1):
        label = total_requests_count - idx + 1
        req.label_id = label
        labeled_requests.append(req)
    device_summary = (
        user_requests
        .values('device__imei_no')
        .annotate(
            total_requested=Count('id'),
            total_approved=Count('id', filter=Q(status='Approved')),
            total_pending=Count('id', filter=Q(status='Pending')),
            total_issued=Count('id', filter=Q(status='Issued')),
            total_fully_returned=Count(
                'id', filter=Q(status='Fully Returned')),
            total_partially_returned=Count(
                'id', filter=Q(status='Partially Returned')),
        )
        .order_by('device__imei_no')
    )
    context = {
        'requests': labeled_requests,
        'total_requests': user_requests.count(),
        'approved_count': user_requests.filter(status='Approved').count(),
        'pending_count': user_requests.filter(status='Pending').count(),
        'issued_count': user_requests.filter(status='Issued').count(),
        'fully_returned_count': user_requests.filter(status='Fully Returned').count(),
        'partially_returned_count': user_requests.filter(status='Partially Returned').count(),
        'device_summary': device_summary,
    }
    return render(request, 'invent/requestor_dashboard.html', context)

# --- Device Request ---
@login_required
def request_device(request):
    user = request.user

    # ... [AJAX handlers remain unchanged] ...

    # =====================================================
    # POST — MULTIPLE DEVICE REQUEST HANDLING
    # =====================================================
    if request.method == "POST":
        client_id = request.POST.get("client_id")
        proof_file = request.FILES.get("payment_proof")

        device_oems = request.POST.getlist("oem[]")
        categories = request.POST.getlist("category[]")
        device_names = request.POST.getlist("device_name[]")
        quantities = request.POST.getlist("quantity[]")

        if not client_id:
            messages.error(request, "Please select a client.")
            return redirect("request_device")

        client = Client.objects.filter(id=client_id).first()
        if not client:
            messages.error(request, "Client not found.")
            return redirect("request_device")

        if not (device_oems and device_names and quantities):
            messages.error(request, "Device request fields are incomplete.")
            return redirect("request_device")

        try:
            with transaction.atomic():

                for idx in range(len(device_oems)):
                    oem_id = device_oems[idx]
                    category_name = categories[idx]
                    device_name = device_names[idx]
                    qty = int(quantities[idx])

                    device = Device.objects.filter(
                        oem_id=oem_id,
                        category=category_name,
                        name=device_name
                    ).first()

                    if not device:
                        messages.error(
                            request, f"Device {device_name} not found.")
                        return redirect("request_device")

                    # IMEI selection
                    imei_filters = Q(device=device, is_available=True)

                    if not user.is_superuser:
                        country = getattr(user.profile, "country", None)
                        if country:
                            imei_filters &= Q(device__branch__country=country)

                    available_imeis = DeviceIMEI.objects.filter(
                        imei_filters).order_by("id")[:qty]

                    if available_imeis.count() < qty:
                        messages.error(
                            request,
                            f"Only {available_imeis.count()} units of {device.name} are available."
                        )
                        return redirect("request_device")

                    # Create DeviceRequest
                    dr = DeviceRequest.objects.create(
                        requestor=user,
                        client=client,
                        device=device,
                        quantity=qty,
                        branch=device.branch,
                        country=device.branch.country,
                        payment_proof=proof_file
                    )

                    # ===========================
                    # Notify store clerks
                    # ===========================
                    store_clerks = User.objects.filter(groups__name="Store Clerk")
                    store_clerk_emails = [u.email for u in store_clerks if u.email]

                    subject = f"New Device Request #{dr.id} Submitted"
                    message = (
                        f"Hello,\n\n"
                        f"A new device request has been submitted by {user.username}.\n"
                        f"Device: {device.name}\n"
                        f"Quantity: {qty}\n"
                        f"Client: {client.name if client else 'N/A'}\n\n"
                        f"Please review and select IMEIs for this request."
                    )

                    if store_clerk_emails:
                        send_mail(
                            subject,
                            message,
                            from_email=None,  # Uses DEFAULT_FROM_EMAIL
                            recipient_list=store_clerk_emails,
                            fail_silently=False,
                        )

                    # Assign available IMEIs (mark as unavailable)
                    for imei_obj in available_imeis:
                        dr.imei_obj = imei_obj
                        dr.imei_no = imei_obj.imei_number
                        dr.save()
                        imei_obj.mark_unavailable()

            messages.success(
                request, "Your device request has been submitted successfully!")
            return redirect("requestor_dashboard")

        except Exception as e:
            messages.error(request, f"Error: {str(e)}")
            return redirect("request_device")

    # =====================================================
    # GET — Load Request Form
    # =====================================================
    return render(request, "invent/request_item.html", {
        "clients": Client.objects.all(),
        "oems": OEM.objects.all(),
    })

# --- Cancel Request ---


@login_required
def cancel_request(request, request_id):
    device_request = get_object_or_404(
        DeviceRequest, id=request_id, requestor=request.user)
    if device_request.status == 'Pending':
        if request.method == 'POST':
            device_request.status = 'Cancelled'
            device_request.save()
            messages.success(
                request, f"Request for '{device_request.device.name}' (ID: {request_id}) has been cancelled.")
            return redirect('requestor_dashboard')
        else:
            context = {
                'device_request': device_request
            }
            return render(request, 'invent/cancel_request_confirm.html', context)
    else:
        messages.error(
            request, f"Request for '{device_request.device.name}' (ID: {request_id}) cannot be cancelled because its status is '{device_request.status}'.")
        return redirect('requestor_dashboard')

# --- Clerk Dashboard ---


@login_required
@permission_required('invent.view_device', raise_exception=True)
def store_clerk_dashboard(request):
    user = request.user
    # COUNTRY FILTER: Restrict by user's country
    if user.is_superuser:
        total_devices = Device.objects.count()
        devices_available = Device.objects.filter(status='available').count()
        devices_issued = Device.objects.filter(status='issued').count()
        devices_returned = Device.objects.filter(status='returned').count()
        recent_issuances = (
            IssuanceRecord.objects
            .select_related('device', 'client')
            .order_by('-issued_at')[:10]
        )
        pending_device_requests = DeviceRequest.objects.filter(
            status="Pending").select_related("requestor", "device", "client")
    else:
        user_country = getattr(user.profile, "country", None)
        total_devices = Device.objects.filter(
            branch__country=user_country).count()
        devices_available = Device.objects.filter(
            status='available', branch__country=user_country).count()
        devices_issued = Device.objects.filter(
            status='issued', branch__country=user_country).count()
        devices_returned = Device.objects.filter(
            status='returned', branch__country=user_country).count()
        recent_issuances = (
            IssuanceRecord.objects
            .filter(device__branch__country=user_country)
            .select_related('device', 'client')
            .order_by('-issued_at')[:10]
        )
        pending_device_requests = DeviceRequest.objects.filter(
            status="Pending", branch__country=user_country).select_related("requestor", "device", "client")
    seen = set()
    recent_devices = []
    for record in recent_issuances:
        if record.device.id not in seen:
            record.device.current_client = record.client
            record.device.issued_at = record.issued_at
            recent_devices.append(record.device)
            seen.add(record.device.id)
        if len(recent_devices) >= 5:
            break
    context = {
        "total_devices": total_devices,
        "devices_available": devices_available,
        "devices_issued": devices_issued,
        "devices_returned": devices_returned,
        "recent_devices": recent_devices,
        "pending_device_requests": pending_device_requests,
    }
    return render(request, 'invent/store_clerk_dashboard.html', context)


@require_POST
def delete_device(request, device_id=None):
    # Permission check
    if not request.user.has_perm('invent.delete_device'):
        messages.error(
            request, "You do not have permission to delete devices.")
        logger.warning(
            "User %s attempted device delete without permission", request.user.username)
        return redirect(reverse('adjust_stock'))

    try:
        if device_id:
            # Single delete
            device = get_object_or_404(Device, pk=device_id)

            # Branch check for non-superusers
            if not request.user.is_superuser:
                user_branch = getattr(request.user.profile, 'branch', None)
                if user_branch is None or device.branch != user_branch:
                    messages.error(
                        request, "You can only delete devices from your assigned branch.")
                    logger.warning(
                        "User %s attempted to delete device %s from another branch", request.user.username, device.id)
                    return redirect(reverse('adjust_stock'))

            device.delete()
            messages.success(
                request, f"Device '{device.name}' deleted successfully.")
            logger.info("User %s deleted device %s",
                        request.user.username, device.id)
        else:
            # Bulk delete: device_ids expected as repeated inputs
            device_ids = request.POST.getlist('device_ids')
            if not device_ids:
                messages.warning(request, "No devices selected for deletion.")
                return redirect(reverse('adjust_stock'))

            qs = Device.objects.filter(id__in=device_ids)

            # Enforce branch scoping for non-superusers
            if not request.user.is_superuser:
                user_branch = getattr(request.user.profile, 'branch', None)
                if user_branch is None:
                    messages.error(
                        request, "Your profile has no branch assigned. Cannot delete devices.")
                    return redirect(reverse('adjust_stock'))
                qs = qs.filter(branch=user_branch)

            deleted_count, _ = qs.delete()
            messages.success(request, f"Deleted {deleted_count} device(s).")
            logger.info("User %s bulk-deleted %d devices",
                        request.user.username, deleted_count)

    except Exception as e:
        logger.exception("Error deleting device(s): %s", e)
        messages.error(request, "An error occurred while deleting devices.")
    return redirect(reverse('adjust_stock'))

# --- Device Request Approval/Reject ---


@login_required
@permission_required('invent.change_devicerequest', raise_exception=True)
def approve_request(request, request_id):
    device_request = get_object_or_404(DeviceRequest, id=request_id)
    device_request.status = "Approved"
    device_request.save()
    messages.success(
        request, f"Request {device_request.id} approved successfully.")
    return redirect("store_clerk_dashboard")


@login_required
@permission_required('invent.change_devicerequest', raise_exception=True)
def reject_request(request, request_id):
    device_request = get_object_or_404(DeviceRequest, id=request_id)
    device_request.status = "Rejected"
    device_request.save()
    messages.warning(request, f"Request {device_request.id} rejected.")
    return redirect("store_clerk_dashboard")

# --- Device Listing/Search ---


@login_required
def inventory_list_view(request):
    query = request.GET.get('q', '')
    status = request.GET.get('status', 'all')
    page = request.GET.get('page', 1)
    per_page = 10  # Set how many devices per page

    user = request.user
    # COUNTRY FILTER: Restrict by user's country
    if user.is_superuser:
        devices = Device.objects.select_related(
            'oem', 'branch').order_by('category', 'oem__name', 'id')
    else:
        user_country = getattr(user.profile, "country", None)
        devices = Device.objects.select_related('oem', 'branch').filter(
            branch__country=user_country).order_by('category', 'oem__name', 'id')

    if status and status != 'all':
        devices = devices.filter(status=status)
    if query:
        devices = devices.filter(
            Q(imei_no__icontains=query) |
            Q(serial_no__icontains=query) |
            Q(name__icontains=query) |
            Q(category__icontains=query) |
            Q(oem__name__icontains=query) |
            Q(issuancerecord__client__name__icontains=query)
        ).distinct()

    for device in devices:
        last_issuance = (
            IssuanceRecord.objects
            .filter(device=device)
            .order_by('-issued_at')
            .select_related('client')
            .first()
        )
        device.current_client = last_issuance.client if last_issuance else None
        device.issued_at = last_issuance.issued_at if last_issuance else None

    grouped_devices = defaultdict(lambda: defaultdict(list))
    for device in devices:
        oem_label = f"{device.oem.name} (ID: {device.oem.id})" if device.oem else "-"
        grouped_devices[device.category][oem_label].append(device)

    paginated_grouped_devices = {}
    for category, oems in grouped_devices.items():
        paginated_grouped_devices[category] = {}
        for oem_label, device_list in oems.items():
            paginator = Paginator(device_list, per_page)
            try:
                page_obj = paginator.page(page)
            except Exception:
                page_obj = paginator.page(1)
            paginated_grouped_devices[category][oem_label] = page_obj

    context = {
        'grouped_devices': paginated_grouped_devices,
        'query': query,
        'status': status,
        'page': page,
    }
    return render(request, 'invent/list_device_grouped.html', context)
# --- Stock Management ---


@login_required
@permission_required('invent.add_device', raise_exception=True)
def manage_stock(request):
    user = request.user
    if request.method == "POST":
        form = DeviceForm(request.POST, user=user)
        if form.is_valid():
            device = form.save(commit=False)
            # Just in case, re-enforce the assignment for non-superusers
            if not user.is_superuser:
                device.branch = user.profile.branch
                device.country = user.profile.country
            device.save()
            messages.success(request, "Device added successfully.")
            return redirect('manage_stock')
        else:
            print(form.errors)  # Debug output for troubleshooting
            messages.error(request, "Please correct the errors below.")
    else:
        form = DeviceForm(user=user)
    return render(request, 'invent/manage_stock.html', {'form': form})


@login_required
@permission_required('invent.change_device', raise_exception=True)
def edit_item(request, device_id):
    device = get_object_or_404(Device, id=device_id)
    if request.method == 'POST':
        form = DeviceForm(request.POST, instance=device)
        if form.is_valid():
            form.save()
            messages.success(
                request, f'Device "{device.name}" updated successfully!')
            return redirect('manage_stock')
        else:
            messages.error(
                request, "Error updating device. Please correct the errors below.")
    else:
        form = DeviceForm(instance=device)
    context = {
        'form': form,
        'device': device,
    }
    return render(request, 'invent/edit_item.html', context)


# Branch Admin Approves/Rejects Selected Devices
@login_required
@permission_required('invent.can_approve_selection', raise_exception=True)
def branch_admin_issue_dashboard(request):
    user = request.user

    # Ensure profile exists
    branch = getattr(user.profile, 'branch', None)
    country = getattr(user.profile, 'country', None)

    # Filter device requests in this branch
    pending_requests = DeviceRequest.objects.filter(
        status__in=['Pending', 'Under Review'],
        branch=branch,
        country=country
    )

    # Filter available devices in the same branch
    available_devices = Device.objects.filter(
        status='available',
        branch=branch,
        country=country
    )

    context = {
        'pending_requests': pending_requests,
        'available_devices': available_devices,
    }

    return render(request, 'invent/branch_admin_issue.html', context)


# --- Issue Device (Clerk) ---

@login_required
@permission_required('invent.can_issue_item', raise_exception=True)
def issue_device(request):
    """
    Handles all device issuance workflows for store clerks.
    """
    user = request.user

    # Filter available data
    available_devices = Device.objects.filter(status='available')
    clients = Client.objects.all()
    pending_requests = DeviceRequest.objects.filter(
        status='Pending').select_related('device', 'client', 'requestor')
    waiting_requests = DeviceRequest.objects.filter(
        status='Waiting Approval').select_related('device', 'client', 'requestor')
    approved_requests = DeviceRequest.objects.filter(
        status='Approved').select_related('device', 'client', 'requestor')
    all_requests = DeviceRequest.objects.all().select_related(
        'device', 'client', 'requestor')

    # -------------------- HANDLE FORM ACTIONS -------------------- #
    if request.method == 'POST':
        action = request.POST.get('action')

        # ---------------- Select IMEIs already handled on another view ---------------- #

        # ----------- DIRECT ISSUE ACTION -----------
        if action == 'direct_issue':
            device_id = request.POST.get('device_id')
            client_id = request.POST.get('client_id')

            if not device_id or not client_id:
                messages.error(
                    request, "Please select both a device and a client.")
                return redirect('issue_device')

            device = get_object_or_404(
                Device, id=device_id, status='available')
            client = get_object_or_404(Client, id=client_id)

            with transaction.atomic():
                # Mark device as issued
                device.status = 'issued'
                device.save()

                # Create issuance record
                IssuanceRecord.objects.create(
                    device=device,
                    client=client,
                    logistics_manager=user,
                    device_request=None  # direct issue has no request
                )

            messages.success(
                request, f"Device {device.device_name} (IMEI: {device.imei_no}) issued directly to {client.name}.")
            return redirect('issue_device')

        # ----------- ISSUE APPROVED REQUEST -----------
        elif action == 'issue':
            request_id = request.POST.get('device_request_id')
            device_request = get_object_or_404(
                DeviceRequest, id=request_id, status='Approved')

            selected_devices = device_request.selected_devices.all()
            if not selected_devices.exists():
                messages.error(
                    request, "No selected IMEIs found for this request.")
                return redirect('issue_device')

            with transaction.atomic():
                for sd in selected_devices:
                    device = sd.device
                    if device.status == 'available':
                        device.status = 'issued'
                        device.save()
                        IssuanceRecord.objects.create(
                            device=device,
                            client=device_request.client,
                            logistics_manager=user,
                            device_request=device_request
                        )

                device_request.status = 'Issued'
                device_request.save()

            messages.success(
                request, f"Devices for Request #{device_request.id} issued successfully.")
            return redirect('issue_device')

        # ----------- INVALID ACTION -----------
        else:
            messages.error(request, "Invalid action.")
            return redirect('issue_device')

    # -------------------- CONTEXT DATA -------------------- #
    context = {
        'available_devices': available_devices,
        'clients': clients,
        'pending_requests': pending_requests,
        'waiting_requests': waiting_requests,
        'approved_requests': approved_requests,
        'all_requests': all_requests,
    }

    return render(request, 'invent/issue_device.html', context)

# --- Select IMEIS ---


from django.contrib.auth.models import User
from django.core.mail import send_mail

@login_required
@permission_required('invent.can_issue_item', raise_exception=True)
def select_imeis(request, request_id):
    device_request = get_object_or_404(DeviceRequest, id=request_id)
    requested_device = device_request.device
    user = request.user

    query = request.GET.get("q", "").strip()

    available_imeis = DeviceIMEI.objects.filter(
        device__name=requested_device.name,
        device__oem=requested_device.oem,
        is_available=True,
    ).select_related('device')

    if query:
        available_imeis = available_imeis.filter(
            Q(imei_number__icontains=query) |
            Q(device__imei_no__icontains=query) |
            Q(device__serial_no__icontains=query) |
            Q(device__name__icontains=query)
        )

    available_count = available_imeis.count()

    if request.method == 'POST':
        created = 0

        # --- Handle Excel Upload ---
        if 'upload_file' in request.FILES:
            excel_file = request.FILES['upload_file']
            try:
                wb = openpyxl.load_workbook(excel_file)
                sheet = wb.active
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    imei_number = str(row[0]).strip()
                    try:
                        imei_obj = DeviceIMEI.objects.get(
                            imei_number=imei_number,
                            device__name=requested_device.name,
                            device__oem=requested_device.oem,
                            is_available=True
                        )
                        SelectedDevice.objects.create(
                            request=device_request,
                            device=imei_obj.device,
                            selected_by=user
                        )
                        imei_obj.is_available = False
                        imei_obj.save(update_fields=['is_available'])
                        created += 1
                    except DeviceIMEI.DoesNotExist:
                        continue

                if created == 0:
                    messages.warning(
                        request, "No valid IMEIs found in the Excel file or all are already assigned.")
                else:
                    messages.success(
                        request, f"{created} IMEI(s) submitted via Excel for Request #{device_request.id}.")

                # Update status
                device_request.status = 'Waiting Approval'
                device_request.save(update_fields=['status'])

                # ====== Notify Admins ======
                admins = User.objects.filter(is_superuser=True)
                admin_emails = [a.email for a in admins if a.email]
                if admin_emails:
                    subject = f"Device Request #{device_request.id} Pending Approval"
                    message = (
                        f"Hello Admin,\n\n"
                        f"The store clerk {user.username} has submitted IMEIs for the following request:\n"
                        f"Device: {device_request.device.name}\n"
                        f"Quantity: {device_request.quantity}\n"
                        f"Requestor: {device_request.requestor.username}\n"
                        f"Client: {device_request.client.name if device_request.client else 'N/A'}\n\n"
                        f"Please review and approve or reject the request."
                    )
                    send_mail(subject, message, from_email=None, recipient_list=admin_emails, fail_silently=False)

                return redirect('issue_device')

            except Exception as e:
                messages.error(request, f"Error reading Excel file: {e}")
                return redirect('select_imeis', request_id=request_id)

        # --- Handle Manual Selection ---
        selected_imei_ids = request.POST.getlist('selected_imeis')
        if selected_imei_ids:
            selected_imeis_qs = DeviceIMEI.objects.filter(
                id__in=selected_imei_ids,
                is_available=True
            ).select_related('device')

            if selected_imeis_qs.exists():
                if selected_imeis_qs.count() > device_request.quantity:
                    messages.warning(
                        request,
                        f"You selected {selected_imeis_qs.count()} IMEIs but the request asked for {device_request.quantity}."
                    )

                for imei_obj in selected_imeis_qs:
                    SelectedDevice.objects.create(
                        request=device_request,
                        device=imei_obj.device,
                        selected_by=user
                    )
                    imei_obj.is_available = False
                    imei_obj.save(update_fields=['is_available'])
                    created += 1

                # Update status
                device_request.status = 'Waiting Approval'
                device_request.save(update_fields=['status'])
                messages.success(
                    request, f"{created} IMEI(s) submitted for Request #{device_request.id}.")

                # ====== Notify Admins ======
                admins = User.objects.filter(is_superuser=True)
                admin_emails = [a.email for a in admins if a.email]
                if admin_emails:
                    subject = f"Device Request #{device_request.id} Pending Approval"
                    message = (
                        f"Hello Admin,\n\n"
                        f"The store clerk {user.username} has submitted IMEIs for the following request:\n"
                        f"Device: {device_request.device.name}\n"
                        f"Quantity: {device_request.quantity}\n"
                        f"Requestor: {device_request.requestor.username}\n"
                        f"Client: {device_request.client.name if device_request.client else 'N/A'}\n\n"
                        f"Please review and approve or reject the request."
                    )
                    send_mail(subject, message, from_email=None, recipient_list=admin_emails, fail_silently=False)

                return redirect('issue_device')

            else:
                messages.error(
                    request, "Selected IMEIs are no longer available.")
                return redirect('select_imeis', request_id=request_id)

        else:
            messages.error(
                request, "Please select at least one IMEI or upload an Excel file.")
            return redirect('select_imeis', request_id=request_id)

    context = {
        'device_request': device_request,
        'available_imeis': available_imeis,
        'query': query,
        'available_count': available_count,
    }
    return render(request, 'invent/select_imeis.html', context)


@login_required
@permission_required('invent.can_issue_item', raise_exception=True)
def submit_devices_for_approval(request):
    """
    Store Clerk posts selected device ids for a given request.
    Creates a DeviceSelectionGroup with status 'Pending'.
    """
    if request.method != 'POST':
        return redirect('issue_device')

    req_id = request.POST.get('request_id')
    # note: select2/multi selects send as list
    selected_ids = request.POST.getlist('device_ids')

    if not req_id or not selected_ids:
        messages.error(request, "Select devices before submitting.")
        return redirect('issue_device')

    device_request = get_object_or_404(DeviceRequest, id=req_id)

    # validate quantity
    requested_qty = device_request.quantity or 1
    if len(selected_ids) != requested_qty:
        messages.error(
            request, f"Please select exactly {requested_qty} devices.")
        return redirect('issue_device')

    # fetch devices and ensure they are available and in same branch/country as clerk
    devices_qs = Device.objects.filter(id__in=selected_ids, status='available')
    if devices_qs.count() != len(selected_ids):
        messages.error(
            request, "One or more selected devices are no longer available.")
        return redirect('issue_device')

    # create selection group
    with transaction.atomic():
        sel_group = DeviceSelectionGroup.objects.create(
            device_request=device_request,
            store_clerk=request.user,
            status='Pending'
        )
        sel_group.devices.set(devices_qs)
        # mark request as under review (optional status)
        device_request.status = 'Under Review'
        device_request.save(update_fields=['status'])

    messages.success(
        request, f"Selected devices submitted for Branch Admin approval (Request {device_request.id}).")
    return redirect('issue_device')


@login_required
@permission_required('invent.can_approve_selection', raise_exception=True)
def approve_device_selection(request):
    """
    Branch Admin approves or rejects a DeviceSelectionGroup.
    - GET: Display all pending selections for review.
    - POST: Approve or reject a specific selection.
    """
    if request.method == 'POST':
        sel_id = request.POST.get('selection_id')
        action = request.POST.get('action')
        sel = get_object_or_404(DeviceSelectionGroup, id=sel_id)

        if action == 'approve':
            with transaction.atomic():
                sel.status = 'Approved'
                sel.reviewed_at = timezone.now()
                sel.reviewed_by = request.user
                sel.save(update_fields=[
                         'status', 'reviewed_at', 'reviewed_by'])

                device_request = sel.device_request
                device_request.status = 'Approved'
                device_request.save(update_fields=['status'])

            messages.success(
                request, f"Selection for Request {device_request.id} approved.")

        elif action == 'reject':
            with transaction.atomic():
                sel.status = 'Rejected'
                sel.reviewed_at = timezone.now()
                sel.reviewed_by = request.user
                sel.save(update_fields=[
                         'status', 'reviewed_at', 'reviewed_by'])

                device_request = sel.device_request
                device_request.status = 'Rejected'
                device_request.save(update_fields=['status'])

            messages.warning(
                request, f"Selection for Request {device_request.id} rejected.")

        else:
            messages.error(request, "Invalid action.")

        # After action, reload the page
        return redirect('approve_device_selection')

    # GET: Show all pending selections for the Branch Admin to review
    pending_selections = (
        DeviceSelectionGroup.objects
        .filter(status='Pending')
        .select_related('device_request', 'store_clerk')
        .prefetch_related('devices')
        .order_by('-created_at')
    )

    context = {
        'pending_selections': pending_selections,
    }

    return render(request, 'invent/Branch_admin_approve.html', context)


@login_required
@permission_required('invent.can_issue_item', raise_exception=True)
def issue_approved_devices(request):
    """
    Store Clerk triggers issuance for an Approved DeviceSelectionGroup (one group per request).
    This will set device.status='issued' and create IssuanceRecord entries.
    """
    if request.method != 'POST':
        return redirect('issue_device')

    req_id = request.POST.get('request_id')
    if not req_id:
        messages.error(request, "Invalid request.")
        return redirect('issue_device')

    device_request = get_object_or_404(DeviceRequest, id=req_id)

    # find the approved selection group for this request (most recent)
    sel = DeviceSelectionGroup.objects.filter(
        device_request=device_request, status='Approved').order_by('-reviewed_at').first()
    if not sel:
        messages.error(
            request, "No approved selection found for this request.")
        return redirect('issue_device')

    client = device_request.client
    if not client:
        messages.error(request, "No client linked to this request.")
        return redirect('issue_device')

    with transaction.atomic():
        for device in sel.devices.select_for_update():  # lock rows
            if device.status != 'available' and device.status != 'reserved':
                messages.error(request, f"Device {device} is not available.")
                return redirect('issue_device')
            device.status = 'issued'
            device.save(update_fields=['status'])
            IssuanceRecord.objects.create(
                device=device,
                client=client,
                logistics_manager=request.user,
                device_request=device_request
            )
        device_request.status = 'Issued'
        device_request.date_issued = timezone.now()
        device_request.save(update_fields=['status', 'date_issued'])

    messages.success(
        request, f"Devices for Request {device_request.id} issued successfully.")
    return redirect('issue_device')

# --- Return Device (Clerk) ---


@login_required
@permission_required('invent.can_issue_item', raise_exception=True)
def return_device(request):
    user = request.user
    # COUNTRY FILTER: Restrict by user's country
    if user.is_superuser:
        issued_devices = Device.objects.filter(status='issued')
    else:
        user_country = getattr(user.profile, "country", None)
        issued_devices = Device.objects.filter(
            status='issued', branch__country=user_country)
    clients = Client.objects.all()
    if request.method == 'POST':
        device_id = request.POST.get('device_id')
        client_id = request.POST.get('client_id')
        reason = request.POST.get('reason', '')
        device = get_object_or_404(Device, id=device_id, status='issued')
        client = get_object_or_404(Client, id=client_id)
        with transaction.atomic():
            device.status = 'returned'
            device.save()
            ReturnRecord.objects.create(
                device=device, client=client, reason=reason)
            messages.success(
                request, f"Device {device.imei_no} returned by {client.name}.")
        return redirect('return_device')
    return render(request, 'invent/return_device.html', {
        'issued_devices': issued_devices,
        'clients': clients
    })

# --- Request Summary for Requestor ---


@login_required
def request_summary(request):
    user_requests = DeviceRequest.objects.filter(requestor=request.user)
    total_requests = user_requests.count()
    pending_requests = user_requests.filter(status='Pending').count()
    approved_requests = user_requests.filter(status='Approved').count()
    issued_requests = user_requests.filter(status='Issued').count()
    rejected_requests = user_requests.filter(status='Rejected').count()
    partially_returned_requests = user_requests.filter(
        status='Partially Returned').count()
    fully_returned_requests = user_requests.filter(
        status='Fully Returned').count()
    total_returned_quantity_by_user = user_requests.aggregate(
        total_returned=Sum('quantity')
    )['total_returned'] or 0
    requests_by_status = user_requests.values(
        'status').annotate(count=Count('id')).order_by('status')
    requests_by_device = user_requests.values('device__name').annotate(
        total_requested=Sum('quantity')
    ).order_by('-total_requested')[:10]
    requests_by_requestor = user_requests.values(
        'requestor__username').annotate(count=Count('id'))
    context = {
        'total_requests': total_requests,
        'pending_requests': pending_requests,
        'approved_requests': approved_requests,
        'issued_requests': issued_requests,
        'rejected_requests': rejected_requests,
        'partially_returned_requests': partially_returned_requests,
        'fully_returned_requests': fully_returned_requests,
        'total_returned_quantity_by_user': total_returned_quantity_by_user,
        'requests_by_status': requests_by_status,
        'requests_by_device': requests_by_device,
        'requests_by_requestor': requests_by_requestor,
    }
    return render(request, 'invent/request_summary.html', context)

# --- Client List ---


@login_required
def client_list(request):
    user = request.user
    query = request.GET.get('q', '')
    status_filter = request.GET.get('status', '')
    client_filter = request.GET.get('client', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    # COUNTRY FILTER: Restrict by user's country
    if user.is_superuser:
        requests_qs = DeviceRequest.objects.select_related(
            'client', 'device').order_by('-date_requested')
    else:
        user_country = getattr(user.profile, "country", None)
        requests_qs = DeviceRequest.objects.select_related(
            'client', 'device').filter(branch__country=user_country).order_by('-date_requested')

    if query:
        requests_qs = requests_qs.filter(
            Q(client_name_icontains=query) |
            Q(client_email_icontains=query) |
            Q(client_phone_no_icontains=query) |
            Q(device_name_icontains=query)
        )

    if status_filter:
        requests_qs = requests_qs.filter(status=status_filter)

    if client_filter:
        requests_qs = requests_qs.filter(client_name_icontains=client_filter)

    if date_from:
        requests_qs = requests_qs.filter(
            date_requested_date_gte=parse_date(date_from))
    if date_to:
        requests_qs = requests_qs.filter(
            date_requested_date_lte=parse_date(date_to))

    paginator = Paginator(requests_qs, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'query': query,
        'status_filter': status_filter,
        'client_filter': client_filter,
        'date_from': date_from,
        'date_to': date_to,
    }
    return render(request, 'invent/client_list.html', context)

# --- Stock Adjustment/Search ---


@login_required
@permission_required('invent.change_device', raise_exception=True)
def adjust_stock(request):
    query = request.GET.get('q', '')
    user = request.user

    # COUNTRY FILTER: Restrict by user's country
    if user.is_superuser:
        devices = Device.objects.all().order_by('id')
    else:
        user_country = getattr(user.profile, "country", None)
        devices = Device.objects.filter(
            branch__country=user_country).order_by('id')

    # Apply search query
    if query:
        devices = devices.filter(
            Q(imei_no__icontains=query) |
            Q(serial_no__icontains=query) |
            Q(name__icontains=query) |
            Q(category__icontains=query) |
            Q(issuancerecord__client__name__icontains=query)
        ).distinct()

    # Annotate with last issuance info
    for device in devices:
        last_issuance = (
            IssuanceRecord.objects
            .filter(device=device)
            .order_by('-issued_at')
            .select_related('client')
            .first()
        )
        device.current_client = last_issuance.client if last_issuance else None
        device.issued_at = last_issuance.issued_at if last_issuance else None

    context = {
        'devices': devices,
        'query': query,
    }
    return render(request, 'invent/adjust_stock.html', context)

# --- OEM Management ---


def add_oem(request):
    edit_mode = False
    oem_to_edit = None

    # Handle OEM deletion
    if request.method == "POST" and "delete_oem_id" in request.POST:
        oem = get_object_or_404(OEM, id=request.POST.get("delete_oem_id"))
        oem.delete()
        messages.success(request, "OEM deleted successfully!")
        return redirect('add_oem')

    # Handle OEM edit (save changes)
    if request.method == "POST" and "edit_oem_id" in request.POST:
        oem_to_edit = get_object_or_404(
            OEM, id=request.POST.get("edit_oem_id"))
        form = OEMForm(request.POST, instance=oem_to_edit)
        edit_mode = True
        if form.is_valid():
            form.save()
            messages.success(request, "OEM updated successfully!")
            return redirect('add_oem')
        else:
            messages.error(request, "Please correct the errors below.")

    # Start editing (GET ?edit=id)
    elif "edit" in request.GET:
        oem_to_edit = get_object_or_404(OEM, id=request.GET.get("edit"))
        form = OEMForm(instance=oem_to_edit)
        edit_mode = True

    # Handle OEM addition
    elif request.method == 'POST':
        form = OEMForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "OEM added successfully!")
            return redirect('add_oem')
        else:
            messages.error(request, "Please correct the errors below.")

    else:
        form = OEMForm()

    oems = OEM.objects.all().order_by('-id')
    return render(request, 'invent/add_oem.html', {
        'form': form,
        'oems': oems,
        'edit_mode': edit_mode,
        'oem_to_edit': oem_to_edit,
    })

# --- Reports and Export ---


@login_required
@permission_required('invent.view_device', raise_exception=True)
def reports_view(request):
    user = request.user
    # COUNTRY FILTER: Restrict by user's country
    if user.is_superuser:
        total_items = Device.objects.aggregate(
            total=Sum('total_quantity'))['total'] or 0
        total_requests = DeviceRequest.objects.count()
        pending_count = DeviceRequest.objects.filter(status='Pending').count()
        approved_count = DeviceRequest.objects.filter(
            status='Approved').count()
        issued_count = DeviceRequest.objects.filter(status='Issued').count()
        rejected_count = DeviceRequest.objects.filter(
            status='Rejected').count()
        fully_returned_count = DeviceRequest.objects.filter(
            status='Fully Returned').count()
        partially_returned_count = DeviceRequest.objects.filter(
            status='Partially Returned').count()
        total_returned_quantity_all_items = DeviceRequest.objects.aggregate(
            total_returned=Sum('returned_quantity')
        )['total_returned'] or 0
        top_requested_items = (
            DeviceRequest.objects.values('device__name')
            .annotate(request_count=Count('id'))
            .order_by('-request_count')[:2]
        )
    else:
        user_country = getattr(user.profile, "country", None)
        total_items = Device.objects.filter(branch__country=user_country).aggregate(
            total=Sum('total_quantity'))['total'] or 0
        total_requests = DeviceRequest.objects.filter(
            branch__country=user_country).count()
        pending_count = DeviceRequest.objects.filter(
            status='Pending', branch__country=user_country).count()
        approved_count = DeviceRequest.objects.filter(
            status='Approved', branch__country=user_country).count()
        issued_count = DeviceRequest.objects.filter(
            status='Issued', branch__country=user_country).count()
        rejected_count = DeviceRequest.objects.filter(
            status='Rejected', branch__country=user_country).count()
        fully_returned_count = DeviceRequest.objects.filter(
            status='Fully Returned', branch__country=user_country).count()
        partially_returned_count = DeviceRequest.objects.filter(
            status='Partially Returned', branch__country=user_country).count()
        total_returned_quantity_all_items = DeviceRequest.objects.filter(branch__country=user_country).aggregate(
            total_returned=Sum('returned_quantity')
        )['total_returned'] or 0
        top_requested_items = (
            DeviceRequest.objects.filter(
                branch__country=user_country).values('device__name')
            .annotate(request_count=Count('id'))
            .order_by('-request_count')[:2]
        )
    context = {
        'total_items': total_items,
        'total_requests': total_requests,
        'pending_count': pending_count,
        'approved_count': approved_count,
        'issued_count': issued_count,
        'rejected_count': rejected_count,
        'fully_returned_count': fully_returned_count,
        'partially_returned_count': partially_returned_count,
        'total_returned_quantity_all_items': total_returned_quantity_all_items,
        'top_requested_items': top_requested_items,
    }
    return render(request, 'invent/reports.html', context)


@login_required
@permission_required('invent.add_device', raise_exception=True)
def upload_inventory(request):
    if request.method == 'POST':
        oem_name = request.POST.get('oem')
        category = request.POST.get('category')
        name = request.POST.get('name')
        excel_file = request.FILES.get('excel_file')

        # Validate basic fields
        if not all([oem_name, category, name]):
            messages.error(
                request, "Please provide OEM, category, and device name before uploading.")
            return redirect("upload_inventory")

        if not excel_file:
            messages.error(request, "Please upload an Excel file.")
            return redirect("upload_inventory")

        # Load Excel
        try:
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active
        except Exception:
            messages.error(
                request, "Invalid Excel file. Please upload a valid .xlsx file.")
            return redirect("upload_inventory")

        header = [str(cell.value).strip().lower()
                  for cell in sheet[1] if cell.value]
        header_index_map = {h: i for i, h in enumerate(header)}

        # Only accept 'imei no' and 'serial no' columns
        if not any(col in header_index_map for col in ["imei no", "serial no"]):
            messages.error(
                request, "Excel file must have at least 'IMEI No' or 'Serial No' column.")
            return redirect("upload_inventory")

        oem_obj, _ = OEM.objects.get_or_create(name=oem_name)

        added_count = 0
        skipped_count = 0

        for row in sheet.iter_rows(min_row=2, values_only=True):
            imei_no = str(row[header_index_map["imei no"]]).strip(
            ) if "imei no" in header_index_map and row[header_index_map["imei no"]] else None
            serial_no = str(row[header_index_map["serial no"]]).strip(
            ) if "serial no" in header_index_map and row[header_index_map["serial no"]] else None

            if not imei_no and not serial_no:
                skipped_count += 1
                continue

            # Skip duplicates
            if imei_no and Device.objects.filter(imei_no=imei_no).exists():
                messages.warning(
                    request, f"IMEI {imei_no} already exists. Skipped.")
                skipped_count += 1
                continue
            if serial_no and Device.objects.filter(serial_no=serial_no).exists():
                messages.warning(
                    request, f"Serial No {serial_no} already exists. Skipped.")
                skipped_count += 1
                continue

            device_kwargs = dict(
                oem=oem_obj,
                name=name,
                category=category,
                imei_no=imei_no or None,
                serial_no=serial_no or None,
                status="available",
            )

            # Assign branch & country if user not superuser
            if not request.user.is_superuser:
                device_kwargs["branch"] = request.user.profile.branch
                device_kwargs["country"] = request.user.profile.country

            Device.objects.create(**device_kwargs)
            added_count += 1

        messages.success(
            request, f"Upload complete. {added_count} devices added, {skipped_count} skipped.")
        return redirect("inventory_list")

    return render(request, "invent/upload_inventory.html")


# --- Total Requests Table/Export (Reports) ---


@login_required
def total_requests(request):
    query = request.GET.get('q', '')
    status_filter = request.GET.get('status', '')
    user = request.user
    # COUNTRY FILTER: Restrict by user's country
    if user.is_superuser:
        queryset = DeviceRequest.objects.select_related(
            "device", "client", "requestor").order_by('-date_requested')
    else:
        user_country = getattr(user.profile, "country", None)
        queryset = DeviceRequest.objects.select_related(
            "device", "client", "requestor").filter(branch__country=user_country).order_by('-date_requested')
    if query:
        queryset = queryset.filter(
            Q(device_imei_no_icontains=query) |
            Q(device_serial_no_icontains=query) |
            Q(device_category_icontains=query) |
            Q(client_name_icontains=query)
        )
    if status_filter and status_filter.lower() != 'all':
        queryset = queryset.filter(status=status_filter)
    paginator = Paginator(queryset, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    context = {
        'page_obj': page_obj,
        'status_filter': status_filter,
    }
    return render(request, 'invent/total_requests.html', context)


@login_required
def export_grouped_inventory(request):
    query = request.GET.get('q', '')
    status = request.GET.get('status', 'all')
    user = request.user

    if user.is_superuser:
        devices = Device.objects.select_related(
            'oem', 'branch').order_by('category', 'oem__name', 'id')
    else:
        user_country = getattr(user.profile, "country", None)
        devices = Device.objects.select_related('oem', 'branch').filter(
            branch__country=user_country).order_by('category', 'oem__name', 'id')

    if status and status != 'all':
        devices = devices.filter(status=status)
    if query:
        devices = devices.filter(
            Q(imei_no__icontains=query) |
            Q(serial_no__icontains=query) |
            Q(name__icontains=query) |
            Q(category__icontains=query) |
            Q(oem__name__icontains=query) |
            Q(oem__oem_id__icontains=query) |
            Q(issuancerecord__client__name__icontains=query)
        ).distinct()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Grouped Inventory"

    headers = ['Category', 'Name', 'OEM', 'OEM ID',
               'IMEI', 'Serial', 'Status', 'Client', 'Issued At']
    ws.append(headers)

    for device in devices:
        last_issuance = (
            IssuanceRecord.objects
            .filter(device=device)
            .order_by('-issued_at')
            .select_related('client')
            .first()
        )
        client_name = last_issuance.client.name if last_issuance and last_issuance.client else "-"
        issued_at = last_issuance.issued_at.strftime(
            '%Y-%m-%d %H:%M') if last_issuance and last_issuance.issued_at else "-"
        ws.append([
            device.category or "-",
            device.name or "-",
            device.oem.name if device.oem else "-",
            device.oem.oem_id if device.oem else "-",
            device.imei_no or "-",
            device.serial_no or "-",
            device.get_status_display() if hasattr(
                device, "get_status_display") else device.status,
            client_name,
            issued_at
        ])

    for i, col in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = 20

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=grouped_inventory_export.xlsx'
    wb.save(response)
    return response


@login_required
def export_total_requests(request):
    status_filter = request.GET.get('status')
    user = request.user
    # COUNTRY FILTER: Restrict by user's country
    if user.is_superuser:
        queryset = DeviceRequest.objects.select_related(
            "device", "client", "requestor")
    else:
        user_country = getattr(user.profile, "country", None)
        queryset = DeviceRequest.objects.select_related(
            "device", "client", "requestor").filter(branch__country=user_country)
    if status_filter:
        queryset = queryset.filter(status=status_filter)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Device Requests"
    headers = [
        "Category", "IMEI", "Serial", "Status", "Client", "Issued At"
    ]
    ws.append(headers)
    for req in queryset:
        device = req.device
        ws.append([
            device.category or "-", device.imei_no or "-", device.serial_no or "-",
            req.status, req.client.name if req.client else "-",
            req.date_issued.strftime(
                "%Y-%m-%d %H:%M") if req.date_issued else "-"
        ])
    for i, col in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = 20
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename=total_requests.xlsx'
    wb.save(response)
    return response


@login_required
def export_inventory_items(request):
    status_filter = request.GET.get('status')
    user = request.user
    # COUNTRY FILTER: Restrict by user's country
    if user.is_superuser:
        queryset = Device.objects.all()
    else:
        user_country = getattr(user.profile, "country", None)
        queryset = Device.objects.filter(branch__country=user_country)
    if status_filter:
        queryset = queryset.filter(status=status_filter)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory Items"
    headers = ['Category', 'IMEI', 'Serial', 'Status', 'Client', 'Issued At']
    ws.append(headers)
    for device in queryset:
        issuance = IssuanceRecord.objects.filter(
            device=device
        ).order_by('-issued_at').first()
        ws.append([
            device.category,
            device.imei_no,
            device.serial_no or "-",
            device.status,
            issuance.client.name if issuance else "-",
            issuance.issued_at.strftime('%Y-%m-%d %H:%M') if issuance else "-"
        ])
    for i, col in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = 20
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=inventory_items.xlsx'
    wb.save(response)
    return response

# --- Return Logic: List of issued requests for return and process return ---


@login_required
@permission_required('invent.can_issue_item', raise_exception=True)
def list_issued_requests_for_return(request):
    user = request.user
    # COUNTRY FILTER: Restrict by user's country
    if user.is_superuser:
        issued_requests = DeviceRequest.objects.filter(
            status='Issued'
        ).select_related('device', 'client', 'requestor').order_by('-date_issued')
    else:
        user_country = getattr(user.profile, "country", None)
        issued_requests = DeviceRequest.objects.filter(
            status='Issued', branch__country=user_country
        ).select_related('device', 'client', 'requestor').order_by('-date_issued')
    context = {
        'issued_requests': issued_requests,
        'title': 'Issued Devices for Return'
    }
    return render(request, 'invent/list_issued_requests_for_return.html', context)


@login_required
@permission_required('invent.can_issue_item', raise_exception=True)
def process_return_for_request(request, request_id):
    device_request = get_object_or_404(
        DeviceRequest, id=request_id, status='Issued'
    )
    if request.method == 'POST':
        returned_quantity = int(request.POST.get('returned_quantity', 1))
        reason = request.POST.get('reason', '')
        if returned_quantity > (device_request.quantity - device_request.returned_quantity):
            messages.error(request, "Cannot return more than what was issued.")
            return redirect('list_issued_requests_for_return')
        with transaction.atomic():
            device_request.returned_quantity += returned_quantity
            if device_request.returned_quantity >= device_request.quantity:
                device_request.status = 'Fully Returned'
            else:
                device_request.status = 'Partially Returned'
            device_request.save()
            device = device_request.device
            device.status = 'returned'
            device.save()
            ReturnRecord.objects.create(
                device=device,
                client=device_request.client,
                reason=reason
            )
            messages.success(
                request, f"{returned_quantity} device(s) marked as returned.")
        return redirect('list_issued_requests_for_return')
    return render(request, 'invent/process_return_for_request.html', {
        'device_request': device_request
    })


@login_required
def request_list(request, status):
    user_requests = DeviceRequest.objects.filter(requestor=request.user)
    if status == "all":
        requests = user_requests
    else:
        requests = user_requests.filter(status__iexact=status)
    return render(request, 'invent/request_list.html', {
        'status': status,
        'requests': requests
    })


@login_required
@permission_required('invent.add_purchaseorder', raise_exception=True)
def purchase_orders(request):
    user = request.user

    # Determine country for filtering (assuming via user profile)
    if user.is_superuser:
        country = None  # Superuser can see all
    else:
        country = getattr(user.profile, 'country', None)
        if not country:
            messages.error(
                request, "No country assigned to your profile. Contact admin.")
            return redirect("store_clerk_dashboard")

    # Handle Add/Edit mode
    edit_mode = False
    po_to_edit = None

    # --- Handle Edit Mode ---
    edit_id = request.GET.get('edit')
    if edit_id:
        po_to_edit = get_object_or_404(PurchaseOrder, id=edit_id)
        # Only allow editing if PO is in user's country
        if not user.is_superuser and po_to_edit.branch.country != country:
            messages.error(
                request, "You do not have permission to edit this purchase order.")
            return redirect('purchase_orders')
        edit_mode = True

    # --- Handle Delete ---
    if request.method == "POST" and 'delete_po_id' in request.POST:
        po_id = request.POST.get("delete_po_id")
        po = get_object_or_404(PurchaseOrder, id=po_id)
        if user.is_superuser or (po.branch.country == country):
            po.delete()
            messages.success(request, "Purchase Order deleted.")
            return redirect('purchase_orders')
        else:
            messages.error(
                request, "You do not have permission to delete this purchase order.")
            return redirect('purchase_orders')

    # --- Handle Form Submission (Add or Edit) ---
    if request.method == "POST" and 'delete_po_id' not in request.POST:
        if edit_mode:
            form = PurchaseOrderForm(
                request.POST, request.FILES, instance=po_to_edit)
        else:
            form = PurchaseOrderForm(request.POST, request.FILES)
        # Restrict branch choices by country for non-superusers
        if not user.is_superuser:
            form.fields['branch'].queryset = Branch.objects.filter(
                country=country)
        # Optionally restrict OEM choices by branch (e.g., only OEMs used in your country's branches)
        # form.fields['oem'].queryset = OEM.objects.all()  # Or filter by another logic if needed
        if form.is_valid():
            po = form.save(commit=False)
            # Set branch.country if needed
            po.save()
            messages.success(
                request, f"Purchase Order {'updated' if edit_mode else 'created'} successfully!")
            return redirect('purchase_orders')
    else:
        if edit_mode:
            form = PurchaseOrderForm(instance=po_to_edit)
        else:
            form = PurchaseOrderForm()
        if not user.is_superuser:
            form.fields['branch'].queryset = Branch.objects.filter(
                country=country)
        # Optionally restrict OEM as above

    # --- List POs for current country ---
    if user.is_superuser:
        purchase_orders = PurchaseOrder.objects.all().order_by('-order_date')
    else:
        purchase_orders = PurchaseOrder.objects.filter(
            branch__country=country).order_by('-order_date')

    return render(request, 'invent/purchase_orders.html', {
        'form': form,
        'purchase_orders': purchase_orders,
        'edit_mode': edit_mode,
        'po_to_edit': po_to_edit,
    })
